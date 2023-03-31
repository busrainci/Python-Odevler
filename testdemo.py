from selenium import webdriver
from webdriver_manager.chrome import ChromeDriverManager
from time import sleep
from selenium.webdriver.common.by import By
from selenium.webdriver.support.wait import WebDriverWait 
from selenium.webdriver.support import expected_conditions as ec
from selenium.webdriver.common.action_chains import ActionChains
import pytest
from pathlib import Path
from datetime import date
import openpyxl
from constants import globalConstants 



#prefix = ön ek test_
# postfix

class DemoClass:
    #her testten önce çalışır
    def setup_method(self):
        self.driver = webdriver.Chrome(ChromeDriverManager().install())
        self.driver.maximize_window()
        self.driver.get(globalConstants.URL)
        self.folderPath = str(date.today())  #24.03.2023
        Path(self.folderPath).mkdir(exist_ok = True)

       
        #günün tarihini al,bu tarih ile bir klasör var mı kontrol et yoksa oluştur.
    #her testten sonra çalışır
    def teardown_method(self):
        self.driver.quit()
    # setup => test_demoFunc -> teardown 
    def test_demoFunc(self):
        #3A Act Arrange Assert
        text = "Hello"
        assert text == "Hello"
    #   setup => test_demo2 -> teardown 
    def test_demo2(self):
        assert True

#   @pytest.mark.skip() # ilgili fonksiyonu çalıştırma demiş olduk
    def getData():
        #veriyi al
        excelFile = openpyxl.load_workbook("data/invalid_login.xls")
        selectedSheet = excelFile["Sayfa1"]

        totalRows = selectedSheet.max_row

        data = []
        for i in range(2,totalRows+1):
            username = selectedSheet.cell(i,1).value
            password = selectedSheet.cell(i,2).value
            tupleData = (username,password)
            data.append(tupleData)
                        
        return data



        return [("1","1"),("kullaniciadim","sifrem"),("kodlamaio","123")]

    @pytest.mark.parametrize("username,password",[("1","1"),("kullanici adim","sifrem")])
    def test_invalid_login(self,username,password):
        self.waitForElementVisible((By.ID,"user-name")) # en fazla 5 saniye olacak şekilde user-name id li elementin görünmesini bekle 
        usernameInput = self.driver.find_element(By.ID,"user-name")
        self.waitForElementVisible((By.ID,"password"),10)
        passwordInput = self.driver.find_element(By.ID,"password")
        sleep(5)
        usernameInput.send_keys(username)
        passwordInput.send_keys(password)
        

        loginBtn = self.driver.find_element(By.ID,"login-button")
        sleep(5)
        
        loginBtn.click()
        sleep(5)
        
        errorMessage = self.driver.find_element(By.XPATH,"//*[@id='login_button_container']/div/form/div[3]/h3")
        self.driver.save_screenshot(f"{self.folderPath}/test-invalid-login-{username}-{password}.png")  #önemliiii !!!!!!!!!!
        #magic string

        assert errorMessage.text == "Epic sadface: Username and password do not match any user in this service"
    
    def waitForElementVisible(self,locator,timeout=5):
        WebDriverWait(self.driver,timeout).until(ec.visibility_of_element_located(locator))
        
      
